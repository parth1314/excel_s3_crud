import json
import openpyxl
import boto3
from io import BytesIO
import logging

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')

def lambda_handler(event, context):
    logger.info(f"Received event: {event}")
    operation = event.get("operation")
    data = event.get("data")
    logger.info(f"Operation: {operation}")
    logger.info(f"Data: {data}")
    logger.info(f"Received event: {json.dumps(event)}")
    bucket_name = 'myexlbucket'
    file_key = 'excelfile.xlsx'

    try:
        if event.get("body"):
            body = json.loads(event["body"])
        else:
            return {
                'statusCode': 400,
                'body': json.dumps({'message': 'Invalid request, no body found'}),
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Access-Control-Allow-Headers': 'Content-Type',
                    'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
                }
            }

        operation = body.get('operation')
        data = body.get('data')

        logger.info(f"Operation: {operation}")
        logger.info(f"Data: {data}")

        logger.info("Downloading the Excel file from S3")
        s3_response = s3.get_object(Bucket=bucket_name, Key=file_key)
        file_content = s3_response['Body'].read()

        logger.info("Loading the Excel file")
        workbook = openpyxl.load_workbook(filename=BytesIO(file_content))
        appendix_c_sheet = workbook['Appendix C-Data']
        appendix_e_sheet = workbook['Appendix E-Data']
        assessment_findings_sheet = workbook['Assessment Findings']

        if operation == 'create':
            if data['appendix'] == 'C':
                last_row = appendix_c_sheet.max_row
                next_serial_number = last_row - 2
                new_row = [next_serial_number] + data['values']
                if len(new_row) < 5:
                    new_row.append("CC-1")  
                appendix_c_sheet.append(new_row)
                logger.info(f"Added row to Appendix C: {new_row}")
            elif data['appendix'] == 'E':
                last_row = appendix_e_sheet.max_row
                next_serial_number = last_row - 2
                new_row = [next_serial_number] + data['values']
                if len(new_row) < 5:
                    new_row.append("CA-1")  
                appendix_e_sheet.append(new_row)
                logger.info(f"Added row to Appendix E: {new_row}")

        elif operation == 'read':
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            rows = []
            for row in sheet.iter_rows(min_row=4, values_only=True):
                row_list = list(row)
                while row_list and row_list[-1] is None:
                    row_list.pop()
                rows.append(row_list)
            return {
                'statusCode': 200,
                'body': json.dumps(rows),
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Access-Control-Allow-Headers': 'Content-Type',
                    'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
                }
            }

        elif operation == 'update':
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            row = int(data['row']) + 4
            col = int(data['col'])  
            value = data['value']
            logger.info('Updating row number: {row} and column number: {col}')
            if row <= 3:
                logger.info('Cannot update header row')
                return

            # Update cell value
            sheet.cell(row=row, column=col).value = value
            logger.info(f"Updated cell ({row}, {col}) to {value}")



        elif operation == 'delete':
            if data['appendix'] == 'C':
                sheet = appendix_c_sheet
            elif data['appendix'] == 'E':
                sheet = appendix_e_sheet

            row = data['row'] + 4  
            logger.info('Deleting row number: {row}')
            if row <= 3:
                logger.info('Cannot delete header row')
                return {
                    'statusCode': 400,
                    'body': json.dumps({'error': 'Cannot delete header row'}),
                    'headers': {
                        'Access-Control-Allow-Origin': '*',
                        'Access-Control-Allow-Headers': 'Content-Type',
                        'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
                    }
                }
            
            
            sheet.delete_rows(row)
            logger.info(f"Deleted row {row}")

            # Shift rows 
            max_row = sheet.max_row

            for r in range(row, max_row):
                for col in range(1, sheet.max_column + 1):
                    if r + 1 <= max_row:
                        sheet.cell(row=r, column=col).value = sheet.cell(row=r + 1, column=col).value
                    else:
                        sheet.cell(row=r, column=col).value = None
            max_row = sheet.max_row
            # Update serial numbers 
            for r in range(4, max_row + 1):  
                sheet.cell(row=r, column=1).value = r - 3 

            logger.info(f"Shifted rows below {row} and updated serial numbers")



        logger.info("Saving the updated Excel file back to S3")
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        s3.put_object(Bucket=bucket_name, Key=file_key, Body=output)
        logger.info(f"File saved back to S3: {file_key}")

        return {
            'statusCode': 200,
            'body': json.dumps({'message': 'Operation successful'}),
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
            }
        }

    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        return {
            'statusCode': 500,
            'body': json.dumps({'message': 'Error', 'error': str(e)}),
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
            }
        }
